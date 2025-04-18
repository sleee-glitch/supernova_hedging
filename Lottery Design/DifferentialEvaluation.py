import numpy as np
import pandas as pd
import numpy as np
from openpyxl import Workbook
import numpy as np
from scipy.optimize import minimize, OptimizeResult
from scipy.optimize import differential_evolution

def generate_HR(scenario):
    rounds = len(scenario['p_values'])
    results = {key: np.zeros(rounds) for key in ['hr_mv_der', 'hr_mv', 'nh_var', 'hr_lpm', 'x1_lpm', 'x2_lpm', 'mean_value', 'mean_value2', 'nh_var1', 'nh_var2', 'skewness_L1', 'skewness_L2']}

    for i in range(rounds):
        p = scenario['p_values'][i]
        x_h = scenario['x1_h_values'][i]
        x_l = scenario['x1_l_values'][i]
        y_h = scenario['x2_h_values'][i]
        y_l = scenario['x2_l_values'][i]

        results['mean_value'][i] = (1 - p) * x_h + p * x_l
        results['mean_value2'][i] = (1 - p) * y_l + p * y_h

        results['nh_var1'][i] = (1 - p) * (x_h - results['mean_value'][i])**2 + p * (x_l - results['mean_value'][i])**2
        results['nh_var2'][i] = p * (y_h - results['mean_value2'][i])**2 + (1 - p) * (y_l - results['mean_value2'][i])**2

        results['skewness_L1'][i] = ((1 - p) * (x_h - results['mean_value'][i])**3 + p * (x_l - results['mean_value'][i])**3) / np.sqrt(results['nh_var1'][i])**3
        results['skewness_L2'][i] = (p * (y_h - results['mean_value2'][i])**3 + (1 - p) * (y_l - results['mean_value2'][i])**3) / np.sqrt(results['nh_var2'][i])**3

        # Differential Evolution for variance optimization
        bounds = [(-100, 100)]  # Adjust bounds as needed
        result_mv = differential_evolution(lambda h: varianceObjective(h[0], p, x_h, x_l, y_h, y_l), bounds)
        results['hr_mv'][i] = result_mv.x[0]

        expected_value = (1 - p) * (x_h + results['hr_mv'][i] * y_l) + p * (x_l + results['hr_mv'][i] * y_h)
        results['nh_var'][i] = (1 - p) * ((x_h + results['hr_mv'][i] * y_l) - expected_value)**2 + p * ((x_l + results['hr_mv'][i] * y_h) - expected_value)**2

        results['hr_mv_der'][i] = np.sqrt(results['nh_var1'][i]) / np.sqrt(results['nh_var2'][i])

        # Differential Evolution for LPM optimization
        result_lpm = differential_evolution(lambda h: lpmObjective(h[0], p, x_h, x_l, y_h, y_l), bounds)
        results['hr_lpm'][i] = result_lpm.x[0]

        results['x1_lpm'][i] = (1 - p) * max(0 - x_h, 0)**2 + p * max(0 - x_l, 0)**2
        results['x2_lpm'][i] = p * max(0 - y_h, 0)**2 + (1 - p) * max(0 - y_l, 0)**2

        # Print debugging information
        print(f"Iteration {i}:")
        print(f"p: {p}, x_h: {x_h}, x_l: {x_l}, y_h: {y_h}, y_l: {y_l}")
        print(f"Best MV: {results['hr_mv'][i]}")
        print(f"Best LPM: {results['hr_lpm'][i]}")
        print("---")

    return results

def varianceObjective(h, p, x_h, x_l, y_h, y_l):
    expected_value = (1 - p) * (x_h + h * y_l) + p * (x_l + h * y_h)
    variance = (1 - p) * ((x_h + h * y_l) - expected_value)**2 + p * ((x_l + h * y_h) - expected_value)**2
    return variance

def lpmObjective(h, p, x_h, x_l, y_h, y_l):
    xbar = 0
    lpm = (1 - p) * max(xbar - (x_h + h * y_l), 0)**2 + p * max(xbar - (x_l + h * y_h), 0)**2
    return lpm





scenarios = {
    'S1': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-6.0, -4.0, -3.05, -2.449, -2.0, -1.633, -1.309, -1.0, -0.667],
        'x1_h_values': [0.7, 1.0, 1.3, 1.6, 2.0, 2.4, 3.1, 4.0, 6.0],
        'x2_h_values': [0.39, 0.7, 1.06, 1.49, 2.0, 2.59, 3.3, 4.3, 6.28],
        'x2_l_values': [-6.28, -4.3, -3.3, -2.59, -2.0, -1.49, -1.06, -0.7, -0.39]
    },
    'S2': {
        'p_values': [0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5, 0.5],
        'x1_l_values': [-8, -7, -6, -5, -4, -3, -2, -1, 0],
        'x1_h_values': [1, 2, 3, 4, 5, 6, 7, 8, 9],
        'x2_h_values': [1, 2, 3, 4, 5, 6, 7, 8, 9],
        'x2_l_values': [-8, -7, -6, -5, -4, -3, -2, -1, 0]
    },
    'S3': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-6.0, -4.0, -3.05, -2.449, -2.0, -1.633, -1.309, -1.0, -0.667],
        'x1_h_values': [0.7, 1.0, 1.3, 1.6, 2.0, 2.4, 3.1, 4.0, 6.0],
        'x2_h_values': [6.4, 4.0, 3.1, 2.4, 2.0, 1.6, 1.3, 1.0, 0.7],
        'x2_l_values': [-0.667, -1.0, -1.309, -1.633, -2.0, -2.449, -3.055, -4.0, -6.0]
    },
    'S4': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4],
        'x1_h_values': [6.3, 4.6, 4.0, 3.7, 3.6, 3.7, 4.0, 4.6, 6.3],
        'x2_h_values': [6.3, 4.6, 4.0, 3.7, 3.6, 3.7, 4.0, 4.6, 6.3],
        'x2_l_values': [-0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4]
    },
    'S5': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-6.0, -4.0, -3.05, -2.449, -2.0, -1.633, -1.309, -1.0, -0.667],
        'x1_h_values': [0.7, 1.0, 1.3, 1.6, 2.0, 2.4, 3.1, 4.0, 6.0],
        'x2_h_values': [6.0, 4.0, 3.06, 2.45, 2.0, 1.63, 1.31, 1.0, 0.67],
        'x2_l_values': [-0.667, -1.0, -1.309, -1.633, -2.0, -2.449, -3.055, -4.0, -6.0]
    },
    'S6': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-6.0, -4.0, -3.05, -2.449, -2.0, -1.633, -1.309, -1.0, -0.667],
        'x1_h_values': [0.7, 1.0, 1.3, 1.6, 2.0, 2.4, 3.1, 4.0, 6.0],
        'x2_h_values': [0.4, 0.4, 0.4, 0.4, 0.4, 0.4, 0.4, 0.4, 0.4],
        'x2_l_values': [-6.3, -4.6, -4.0, -3.7, -3.6, -3.7, -4.0, -4.6, -6.3]
    },
    'S7': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-6.0, -4.0, -3.05, -2.449, -2.0, -1.633, -1.309, -1.0, -0.667],
        'x1_h_values': [0.7, 1.0, 1.3, 1.6, 2.0, 2.4, 3.1, 4.0, 6.0],
        'x2_h_values': [6.3, 4.6, 4.0, 3.7, 3.6, 3.7, 4.0, 4.6, 6.3],
        'x2_l_values': [-0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4]
    },
    'S8': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-6.0, -4.0, -3.055, -2.449, -2.0, -1.633, -1.309, -1.0, -0.667],
        'x1_h_values': [0.7, 1.0, 1.3, 1.6, 2.0, 2.4, 3.1, 4.0, 6.0],
        'x2_h_values': [10.00, 7.40, 6.40, 5.95, 5.83, 5.95, 6.40, 7.40, 10.00],
        'x2_l_values': [-0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4, -0.4]
    },
    'S9': {
        'p_values': [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9],
        'x1_l_values': [-6.0, -4.0, -3.05, -2.449, -2.0, -1.633, -1.309, -1.0, -0.667],
        'x1_h_values': [0.7, 1.0, 1.3, 1.6, 2.0, 2.4, 3.1, 4.0, 6.0],
        'x2_h_values': [0.4, 0.4, 0.4, 0.4, 0.4, 0.4, 0.4, 0.4, 0.4],
        'x2_l_values': [-10.00, -7.40, -6.40, -5.95, -5.83, -5.95, -6.40, -7.40, -10.00]
    }
}




def collect_scenario_results(scenario, scenario_num):
    hr = generate_HR(scenario)
    rounds = len(scenario['p_values'])

    df = pd.DataFrame({
        '1-p': 1 - np.array(scenario['p_values']),
        'L1_high': scenario['x1_h_values'],
        'p': scenario['p_values'],
        'L1_low': scenario['x1_l_values'],
        'L1_mean': hr['mean_value'],
        'L1_variance': hr['nh_var1'],
        'L1_skewness': hr['skewness_L1'],
        'L2_high': scenario['x2_h_values'],
        'L2_low': scenario['x2_l_values'],
        'L2_mean': hr['mean_value2'],
        'L2_variance': hr['nh_var2'],
        'L2_skewness': hr['skewness_L2'],
        'hr_mv': hr['hr_mv'],
        'hr_lpm': hr['hr_lpm']
    })

    return df


wb = Workbook()

# Loop through scenarios and add each to the workbook
for i, scenario in enumerate(scenarios.values(), 1):
    scenario_results = collect_scenario_results(scenario, i)
    sheet = wb.create_sheet(title=f"Scenario {i}")

    # Write the column headers
    for col, header in enumerate(scenario_results.columns, 1):
        sheet.cell(row=1, column=col, value=header)

    # Write the data
    for row, data in enumerate(scenario_results.values, 2):
        for col, value in enumerate(data, 1):
            sheet.cell(row=row, column=col, value=value)

# Remove the default sheet created by openpyxl
wb.remove(wb['Sheet'])

# Save the workbook to an Excel file
wb.save("HR_Scenarios7.xlsx")

